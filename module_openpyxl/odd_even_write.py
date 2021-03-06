"""
# 엑셀작업 줄이기 - 엑셀없이 코드로 수정\\정리\\추출\\쓰기
# OpenPyXL = xlsx 지원, 읽기\\쓰기
# XlsxWriter (화일생성, 서버사이드) = xlsx 지원
# Xlrd, Xlwt, xlwings, pywin32 = xls 지원 (Win, Mac, Linux)
# xlwings (W,M), pywin32 (W) - 설치 된 엑셀필요
"""
# script run 에서는 실행이 안됨.

import pandas as pd
import openpyxl as opx

import assets.script_run
from assets.configs import *
from assets.classes import *


# 같은 내용 다른 포맷
TARGET_READ = dir_home_statics + '_result.xlsx'
OUTPUT_01 = dir_home_works + '_u_even.xls'
OUTPUT_02 = dir_home_works + '_u_odd.xls'

WB = opx.load_workbook(TARGET_READ)
WS = WB["Sheet"]
# sheet = wb.get_sheet_by_name("Sheet1")        # deprecated


def excel_read_split_out(read, out1, out2):
    """
    # TARGET 을 읽어 들여서 ... 판다스로 가공
    # pd.df.to_xls() output_01, 02 로 나누어 저장.
    # DOCS::pandas.read_excel = https:/goo.gl/Dy1A5z
    """
    # read target file as pandas df
    df = pd.read_excel(read)
    print(df)           # for TEST

    # rows_even
    df[0::2].to_excel(out1)

    # rows_odd
    df[1::2].to_excel(out2)

def show_cells(sheet):
    """
    # 셀내용 보기 - print(sheet['A1'].value, sheet['B1'].value)
    """
    # 셀 값 정보조회 하기
    print(f"  - Active cell(now): '{sheet.active_cell}'")
    print(f"  - min/max rows    : {sheet.min_row} ~ {sheet.max_row}")
    print(f"  - min/max columns : {sheet.min_column} ~ {sheet.max_column}")
    print(f"  - active range    : {sheet.calculate_dimension()}")
    print()

    # A, B 두줄을 끝(1~10)까지 조회 한다.
    for i in range(1, sheet.max_row + 1):
        cell_01 = sheet[f"A{i}"].value
        cell_02 = sheet[f"B{i}"].value
        print(f"{cell_01:>10} : {cell_02}")

    # row, col = xy로 값(value) 조회가능
    print("\n\n*** row, col = xy로 값(value) 조회가능 ***")
    print(f"[1,1] = {sheet.cell(1, 1).value}")
    print(f"[1,2] = {sheet.cell(1, 2).value}")

def show_overview(sheet):
    """
    # 전체 입력 란을 오버뷰
    """
    for row in sheet.rows:
        print(type(row[5]))   # <class 'openpyxl.cell.cell.Cell'>

        # [1, 2, 3, 4, 5, 6]
        print([col.col_idx for col in row])

        # ['NAME', 'KOR', 'ENG', 'MATH', 'SUM', 'AVERAGE']
        print([col.value for col in row])

        # 뒤에서 3번째 칼럼 = MATH
        print(row[-3].value)
        pass
    # [print(func) for func in dir(row[5]) if not func.startswith('_')]

def run01():
    print(excel_read_split_out.__doc__)
    excel_read_split_out(TARGET_READ, OUTPUT_01, OUTPUT_02)

def run02():
    print(show_cells.__doc__)
    show_cells(WS)

def run03():
    print(show_overview.__doc__)
    show_overview(WS)

def main():
    # run01()
    # run02()
    run03()


if __name__ == '__main__':
    main()
