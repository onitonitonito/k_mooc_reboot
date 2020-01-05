"""
* 엑셀작업 줄이기 - 엑셀없이 코드로 수정/정리/추출/쓰기
파이썬으로 엑셀 보고서 만들기 = http://www.hanul93.com/openpyxl-basic/
"""
# script run 에서는 실행이 안됨.

import openpyxl as opx
import assets.script_run

from assets.configs import dir_home_statics

filename_read = dir_home_statics + '_write.xlsx'
filename_write = dir_home_statics + '_result.xlsx'

# 엑셀파일 열기 (WB = Working Book)
WB = opx.load_workbook(filename_read)
WS = WB.active

# 국영수 점수를 읽기
for row in WS.rows:
    row_index = row[0].row   # 현재 행 인덱스 1~
    name = row[0].value
    kor = row[1].value
    eng = row[2].value
    math = row[3].value

    # kor 위치의 값을 기준으로, 판단하여 sum, average 변경
    if isinstance(kor, str):
        (sum, average) = "SUM", "AVERAGE"

    elif isinstance(kor, int):
        sum = f"=SUM(B{row_index}:D{row_index})"
        average = f"=ROUND(AVERAGE(B{row_index}:D{row_index}), 0)"

    else:
        sum = average = "... NoneType!!"

    # 셀에 합계 쓰기 ... row_index (현재행)
    WS.cell(row=row_index, column=5).value = sum
    WS.cell(row=row_index, column=6).value = average

    print(f"row_num = {row_index} ... {name}, {kor}, {eng}, {math}, {sum}, {average}")

# 마지막 라인, 최종 합계란 생성
WS.append(['TOTAL SUM',
           '=SUM(B2:B9)',
           '=SUM(C2:C9)',
           '=SUM(D2:D9)',
           '=ROUND(AVERAGE(E2:E9), 0)',
           '=ROUND(AVERAGE(F2:F9), 0)',
           ])

# 엑셀 파일 저장
WB.save(filename_write)
WB.close()
