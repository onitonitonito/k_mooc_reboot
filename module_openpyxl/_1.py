
import os
from openpyxl import Workbook

DESTIN_DIR = os.path.join(os.path.dirname(__file__),'')
print(DESTIN_DIR)

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 'Python OpenPyXL module TEST'

# Python types will automatically be converted
import datetime
ws['B1'] = datetime.datetime.now()

# Rows can also be appended
ws.append([1, 2, 3])

for content in ws['A1'].__dir__():
    print(content)

print(ws['A1'].value)       # 42
print(ws['B1'].value)       # datetime.now()

# Save the file
wb.save(DESTIN_DIR+ "_sample.xlsx")
